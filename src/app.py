
from flask import Flask,render_template,request
from flask_mysqldb import MySQL
from openpyxl import load_workbook
import os.path
app = Flask(__name__)
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = '********'
app.config['FILE_UPLOADS']=r"D:\PycharmProjects\dataconverter\src\static\upload"
mysql = MySQL(app)
@app.route('/', methods=["GET"])
def index():
    db = [] 
    cur = mysql.connection.cursor()
    cur.execute(f'SHOW DATABASES;')
    show = cur.fetchall()
    for sh in show:
        for s in sh:
            db.append(s)
    return render_template('index.html', db=db )    
@app.route('/success',methods=["GET", "POST"])
def success():
    tb = []
    vallist = []
    tbcol = []
    if request.method == 'POST':
        database = request.form.get("database")
        table = request.form.get("table")
        spreadsheet = request.form.get("spreadsheet")
        col_type = request.form.get("col_type")
        rows = request.form.get("rows")
        col = request.form.get("col") 
        if request.files:
            excel = request.files['excel']
            excel.save(os.path.join(app.config['FILE_UPLOADS'], excel.filename))       
            wrkbk = load_workbook(excel)
            hj = wrkbk[f"{spreadsheet}"]
            
            cur = mysql.connection.cursor()
            cur.execute(f"USE {database}")
            cur.execute("SHOW TABLES")
            sho = cur.fetchall()
            for sh in sho:
              for s in sh:
                tb.append(s)
            if table in tb:
                print("already exists")
            else:
                  tbcol.append(col_type)
                  a = (",".join(tbcol))
                  corr = f"({a})"
                  rw=int(rows)
                  cl=int(col)
                  cur.execute(f"CREATE TABLE {table} {corr};")
                  print("hh")
                  for row in hj.iter_rows(min_row=2, max_row=rw, min_col=1, max_col=cl, values_only=True):
                    vallist.append(row)
                  qw = str(vallist)[1:-1]
                  cur.execute(f"INSERT INTO {table} VALUES {qw};")
                  mysql.connection.commit()
                  print("p")
         
    return render_template('success.html')
if __name__ == '__main__':
    app.run()
